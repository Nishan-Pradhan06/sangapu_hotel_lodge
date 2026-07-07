import io
from decimal import Decimal
from datetime import datetime

from django.http import HttpResponse
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated

from rooms.models import RoomEntry
from expenses.models import ExpenseEntry


# ─────────────────────────────────────────────
#  Shared helpers: build + filter records
# ─────────────────────────────────────────────
def _build_income_records(entries):
    records = []
    for entry in entries:
        price = entry.fixed_price if entry.fixed_price is not None else entry.custom_price
        records.append({
            'id': entry.id,
            'type': 'Income',
            'category': 'Fixed Price' if entry.fixed_price is not None else 'Custom Price',
            'amount': float(price),
            'remarks': entry.additional_notes or '',
            'nepali_date': entry.nepali_date,
            'created_at': entry.created_at,
        })
    return records


def _build_expense_records(entries):
    records = []
    for entry in entries:
        records.append({
            'id': entry.id,
            'type': 'Expense',
            'category': entry.get_category_display(),
            'amount': float(entry.amount),
            'remarks': entry.remarks or '',
            'nepali_date': entry.nepali_date,
            'created_at': entry.created_at,
        })
    return records


def get_filtered_records(request):
    """Apply query params and return (records, total_income, total_expenses)."""
    filter_type  = request.query_params.get('type', 'all')
    filter_date  = request.query_params.get('date', None)
    filter_month = request.query_params.get('month', None)
    ordering     = request.query_params.get('ordering', 'desc')
    income_qs  = RoomEntry.objects.filter(user=request.user)
    expense_qs = ExpenseEntry.objects.filter(user=request.user)

    if filter_date:
        income_qs  = income_qs.filter(nepali_date=filter_date)
        expense_qs = expense_qs.filter(nepali_date=filter_date)
    if filter_month:
        income_qs  = income_qs.filter(nepali_date__startswith=filter_month)
        expense_qs = expense_qs.filter(nepali_date__startswith=filter_month)

    records        = []
    total_income   = Decimal('0')
    total_expenses = Decimal('0')

    if filter_type in ('income', 'all'):
        inc = _build_income_records(income_qs)
        records.extend(inc)
        for r in inc:
            total_income += Decimal(str(r['amount']))

    if filter_type in ('expenses', 'all'):
        exp = _build_expense_records(expense_qs)
        records.extend(exp)
        for r in exp:
            total_expenses += Decimal(str(r['amount']))

    reverse = (ordering == 'desc')
    records.sort(key=lambda x: (x['nepali_date'], str(x['created_at'])), reverse=reverse)

    return records, total_income, total_expenses


# ─────────────────────────────────────────────
#  JSON Statement
# ─────────────────────────────────────────────
class StatementView(APIView):
    """
    Bank-style statement combining all income and expenses.

    Query Params: type, date, month, ordering
    """
    permission_classes = [IsAuthenticated]

    def get(self, request, *args, **kwargs):
        records, total_income, total_expenses = get_filtered_records(request)
        net_balance = total_income - total_expenses

        return Response({
            'summary': {
                'total_income':   float(total_income),
                'total_expenses': float(total_expenses),
                'net_balance':    float(net_balance),
                'total_records':  len(records),
            },
            'transactions': records,
        })


# ─────────────────────────────────────────────
#  Excel Export  — Bank Statement Level
# ─────────────────────────────────────────────
class StatementExcelExportView(APIView):
    """
    Export statement as Excel (.xlsx) — A4 portrait, bank-statement design.

    Query Params: type, date, month, ordering
    """
    permission_classes = [IsAuthenticated]

    def get(self, request, *args, **kwargs):
        from openpyxl import Workbook
        from openpyxl.styles import (
            Font, PatternFill, Alignment, Border, Side, GradientFill
        )
        from openpyxl.utils import get_column_letter
        from openpyxl.worksheet.page import PageMargins

        records, total_income, total_expenses = get_filtered_records(request)
        net_balance = total_income - total_expenses
        generated_at = datetime.now().strftime("%Y-%m-%d  %H:%M:%S")
        user_name = getattr(request.user, 'name', '') or request.user.email

        wb = Workbook()
        ws = wb.active
        ws.title = "Bank Statement"

        # ── A4 portrait page setup ──────────────────────────────────────────
        ws.page_setup.paperSize   = ws.PAPERSIZE_A4        # A4
        ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
        ws.page_setup.fitToPage   = True
        ws.page_setup.fitToWidth  = 1
        ws.page_setup.fitToHeight = 0
        ws.page_margins = PageMargins(
            left=0.5, right=0.5, top=0.75, bottom=0.75,
            header=0.3, footer=0.3
        )
        ws.print_title_rows = '7:7'   # repeat column-header row on each printed page

        # ── Colour / style constants ────────────────────────────────────────
        C_NAVY      = "1A3C5E"   # deep navy — header bg
        C_BLUE      = "2563EB"   # accent blue
        C_LIGHT     = "EFF6FF"   # very light blue tint for alt rows
        C_WHITE     = "FFFFFF"
        C_INCOME    = "D1FAE5"   # soft green
        C_INCOME_TXT= "065F46"
        C_EXPENSE   = "FEE2E2"   # soft red
        C_EXPENSE_TXT="991B1B"
        C_GOLD      = "FEF9C3"   # summary highlight
        C_BALANCE   = "DBEAFE"   # net-balance row
        C_GREY_TXT  = "6B7280"
        C_DARK_TXT  = "111827"
        C_DIVIDER   = "BFDBFE"

        def solid(hex_color):
            return PatternFill("solid", fgColor=hex_color)

        def border_side(style="thin", color="D1D5DB"):
            return Side(style=style, color=color)

        thin_border = Border(
            left=border_side(), right=border_side(),
            top=border_side(), bottom=border_side()
        )
        thick_bottom = Border(
            left=border_side(), right=border_side(),
            top=border_side(), bottom=border_side("medium", "1A3C5E")
        )
        header_border = Border(
            left=border_side("medium", C_NAVY),
            right=border_side("medium", C_NAVY),
            top=border_side("medium", C_NAVY),
            bottom=border_side("medium", C_NAVY),
        )

        def set_font(cell, bold=False, size=10, color=C_DARK_TXT, italic=False, name="Calibri"):
            cell.font = Font(name=name, bold=bold, size=size, color=color, italic=italic)

        def center(cell):
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        def vcenter(cell):
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

        def right_align(cell):
            cell.alignment = Alignment(horizontal="right", vertical="center")

        # ── Column widths (A4 portrait = ~120 units across 7 cols) ─────────
        col_widths = [5, 13, 14, 14, 22, 16, 36]
        for i, w in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

        row = 1

        # ══════════════════════════════════════════════════════════════════
        # ROW 1-2  BANK LOGO / HEADER BANNER
        # ══════════════════════════════════════════════════════════════════
        ws.merge_cells(f"A{row}:G{row}")
        c = ws.cell(row=row, column=1, value="🏨  HOTEL LODGE")
        c.font      = Font(name="Calibri", bold=True, size=18, color=C_WHITE)
        c.fill      = solid(C_NAVY)
        c.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[row].height = 34
        row += 1

        ws.merge_cells(f"A{row}:G{row}")
        c = ws.cell(row=row, column=1, value="OFFICIAL ACCOUNT STATEMENT")
        c.font      = Font(name="Calibri", bold=False, size=10, color="93C5FD", italic=True)
        c.fill      = solid(C_NAVY)
        c.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[row].height = 18
        row += 1

        # ── thin separator ─────────────────────────────────────────────────
        ws.merge_cells(f"A{row}:G{row}")
        c = ws.cell(row=row, column=1, value="")
        c.fill = solid(C_BLUE)
        ws.row_dimensions[row].height = 3
        row += 1

        # ══════════════════════════════════════════════════════════════════
        # ROW 4  Meta info  (Account Holder | Generated)
        # ══════════════════════════════════════════════════════════════════
        ws.row_dimensions[row].height = 20
        # left: account holder
        ws.merge_cells(f"A{row}:D{row}")
        c = ws.cell(row=row, column=1, value=f"Account Holder:  {user_name}")
        set_font(c, bold=True, size=10, color=C_NAVY)
        vcenter(c)
        c.fill = solid("F0F9FF")
        # right: generated date
        ws.merge_cells(f"E{row}:G{row}")
        c = ws.cell(row=row, column=5, value=f"Generated:  {generated_at}")
        set_font(c, size=9, color=C_GREY_TXT, italic=True)
        right_align(c)
        c.fill = solid("F0F9FF")
        row += 1

        # ══════════════════════════════════════════════════════════════════
        # ROW 5  Summary Cards  (Total Income | Total Expenses | Net Balance)
        # ══════════════════════════════════════════════════════════════════
        ws.row_dimensions[row].height = 28

        # Card 1 – Income
        ws.merge_cells(f"A{row}:B{row}")
        c = ws.cell(row=row, column=1,
                    value=f"▲ Total Income\nRs. {float(total_income):,.2f}")
        c.font      = Font(name="Calibri", bold=True, size=10, color=C_INCOME_TXT)
        c.fill      = solid(C_INCOME)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border    = thin_border
        ws.row_dimensions[row].height = 34

        # Card 2 – Expenses
        ws.merge_cells(f"C{row}:D{row}")
        c = ws.cell(row=row, column=3,
                    value=f"▼ Total Expenses\nRs. {float(total_expenses):,.2f}")
        c.font      = Font(name="Calibri", bold=True, size=10, color=C_EXPENSE_TXT)
        c.fill      = solid(C_EXPENSE)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border    = thin_border

        # Card 3 – Net Balance
        bal_color = "065F46" if net_balance >= 0 else "991B1B"
        bal_bg    = "D1FAE5" if net_balance >= 0 else "FEE2E2"
        bal_sign  = "+" if net_balance >= 0 else ""
        ws.merge_cells(f"E{row}:G{row}")
        c = ws.cell(row=row, column=5,
                    value=f"◆ Net Balance\nRs. {bal_sign}{float(net_balance):,.2f}")
        c.font      = Font(name="Calibri", bold=True, size=11, color=bal_color)
        c.fill      = solid(bal_bg)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border    = Border(
            left=border_side("medium", C_BLUE), right=border_side("medium", C_BLUE),
            top=border_side("medium", C_BLUE),  bottom=border_side("medium", C_BLUE),
        )
        row += 1

        # ── divider ────────────────────────────────────────────────────────
        ws.merge_cells(f"A{row}:G{row}")
        c = ws.cell(row=row, column=1, value="")
        c.fill = solid(C_DIVIDER)
        ws.row_dimensions[row].height = 3
        row += 1

        # ══════════════════════════════════════════════════════════════════
        # ROW 7  Column Headers
        # ══════════════════════════════════════════════════════════════════
        headers = ["#", "Date (BS)", "Type", "Category", "Amount (Rs.)", "Recorded At", "Remarks"]
        ws.row_dimensions[row].height = 22
        for col, h in enumerate(headers, 1):
            c = ws.cell(row=row, column=col, value=h)
            c.font      = Font(name="Calibri", bold=True, size=10, color=C_WHITE)
            c.fill      = solid(C_NAVY)
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.border    = header_border
        row += 1

        # ══════════════════════════════════════════════════════════════════
        # DATA ROWS
        # ══════════════════════════════════════════════════════════════════
        for idx, r in enumerate(records, 1):
            is_income = r['type'] == 'Income'
            row_fill  = solid(C_INCOME if is_income else C_EXPENSE)
            alt_fill  = solid("F0FDF4" if is_income else "FFF5F5")
            use_fill  = row_fill if idx % 2 == 1 else alt_fill
            txt_color = C_INCOME_TXT if is_income else C_EXPENSE_TXT
            ws.row_dimensions[row].height = 18

            values = [
                idx,
                r['nepali_date'],
                r['type'],
                r['category'],
                r['amount'],
                str(r['created_at'])[:19],
                r['remarks'],
            ]
            for col, val in enumerate(values, 1):
                c = ws.cell(row=row, column=col, value=val)
                c.fill   = use_fill
                c.border = thin_border
                c.alignment = Alignment(vertical="center", wrap_text=(col == 7))

                if col == 1:    # #
                    c.font      = Font(name="Calibri", size=9, color=C_GREY_TXT)
                    c.alignment = Alignment(horizontal="center", vertical="center")
                elif col == 2:  # date
                    c.font      = Font(name="Calibri", size=9, bold=True, color=C_DARK_TXT)
                elif col == 3:  # type
                    c.font      = Font(name="Calibri", size=9, bold=True, color=txt_color)
                    c.alignment = Alignment(horizontal="center", vertical="center")
                elif col == 4:  # category
                    c.font      = Font(name="Calibri", size=9, color=C_DARK_TXT)
                    c.alignment = Alignment(horizontal="center", vertical="center")
                elif col == 5:  # amount
                    c.font          = Font(name="Calibri", size=10, bold=True, color=txt_color)
                    c.number_format = '#,##0.00'
                    c.alignment     = Alignment(horizontal="right", vertical="center")
                elif col == 6:  # recorded at
                    c.font      = Font(name="Calibri", size=8, color=C_GREY_TXT, italic=True)
                    c.alignment = Alignment(horizontal="center", vertical="center")
                else:           # remarks
                    c.font      = Font(name="Calibri", size=9, color=C_DARK_TXT)
            row += 1

        # ══════════════════════════════════════════════════════════════════
        # TOTALS FOOTER ROW
        # ══════════════════════════════════════════════════════════════════
        ws.row_dimensions[row].height = 22
        ws.merge_cells(f"A{row}:D{row}")
        c = ws.cell(row=row, column=1, value=f"TOTAL  ({len(records)} transactions)")
        c.font      = Font(name="Calibri", bold=True, size=10, color=C_WHITE)
        c.fill      = solid(C_NAVY)
        c.alignment = Alignment(horizontal="right", vertical="center")
        c.border    = thick_bottom

        c = ws.cell(row=row, column=5, value=float(total_income - total_expenses))
        c.font          = Font(name="Calibri", bold=True, size=11, color=C_WHITE)
        c.fill          = solid(C_BLUE)
        c.number_format = '#,##0.00'
        c.alignment     = Alignment(horizontal="right", vertical="center")
        c.border        = thick_bottom

        for col in [6, 7]:
            c = ws.cell(row=row, column=col, value="")
            c.fill   = solid(C_NAVY)
            c.border = thick_bottom
        row += 1

        # ── footer note ────────────────────────────────────────────────────
        ws.row_dimensions[row].height = 14
        ws.merge_cells(f"A{row}:G{row}")
        c = ws.cell(row=row, column=1,
                    value="This is a system-generated statement. No signature required.")
        c.font      = Font(name="Calibri", size=8, color=C_GREY_TXT, italic=True)
        c.fill      = solid("F8FAFC")
        c.alignment = Alignment(horizontal="center", vertical="center")

        # ── Stream ─────────────────────────────────────────────────────────
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        response = HttpResponse(
            buffer,
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = 'attachment; filename="hotel_lodge_statement.xlsx"'
        return response


# ─────────────────────────────────────────────
#  PDF Export  — Bank Statement Level (A4 Portrait)
# ─────────────────────────────────────────────
class StatementPDFExportView(APIView):
    """
    Export statement as PDF — A4 portrait, bank-statement design.

    Query Params: type, date, month, ordering
    """
    permission_classes = [IsAuthenticated]

    def get(self, request, *args, **kwargs):
        from reportlab.lib.pagesizes import A4
        from reportlab.lib import colors
        from reportlab.lib.units import mm
        from reportlab.platypus import (
            SimpleDocTemplate, Table, TableStyle, Paragraph,
            Spacer, HRFlowable, KeepTogether
        )
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.enums import TA_CENTER, TA_RIGHT, TA_LEFT
        from reportlab.platypus import Image
        from reportlab.graphics.shapes import Drawing, Rect
        from reportlab.graphics import renderPDF

        records, total_income, total_expenses = get_filtered_records(request)
        net_balance   = total_income - total_expenses
        generated_at  = datetime.now().strftime("%Y-%m-%d  %H:%M:%S")
        user_name     = getattr(request.user, 'name', '') or request.user.email
        total_records = len(records)

        # ── Colours ────────────────────────────────────────────────────────
        NAVY        = colors.HexColor('#1A3C5E')
        BLUE        = colors.HexColor('#2563EB')
        LIGHT_BLUE  = colors.HexColor('#DBEAFE')
        INCOME_BG   = colors.HexColor('#D1FAE5')
        INCOME_TXT  = colors.HexColor('#065F46')
        EXPENSE_BG  = colors.HexColor('#FEE2E2')
        EXPENSE_TXT = colors.HexColor('#991B1B')
        GOLD        = colors.HexColor('#FEF9C3')
        WHITE       = colors.white
        GREY        = colors.HexColor('#6B7280')
        DARK        = colors.HexColor('#111827')
        DIVIDER     = colors.HexColor('#BFDBFE')
        ALT_ROW     = colors.HexColor('#F8FAFC')

        bal_bg  = INCOME_BG if net_balance >= 0 else EXPENSE_BG
        bal_txt = INCOME_TXT if net_balance >= 0 else EXPENSE_TXT

        # ── Page setup (A4 portrait) ────────────────────────────────────────
        buffer = io.BytesIO()
        PAGE_W, PAGE_H = A4
        MARGIN = 18 * mm
        CONTENT_W = PAGE_W - 2 * MARGIN

        doc = SimpleDocTemplate(
            buffer,
            pagesize=A4,
            rightMargin=MARGIN, leftMargin=MARGIN,
            topMargin=MARGIN,   bottomMargin=MARGIN,
        )

        # ── Paragraph styles ───────────────────────────────────────────────
        def make_style(name, **kw):
            base = getSampleStyleSheet()['Normal']
            return ParagraphStyle(name, parent=base, **kw)

        header_title  = make_style('HTitle',  fontSize=20, textColor=WHITE,    alignment=TA_CENTER, fontName='Helvetica-Bold', spaceAfter=0, spaceBefore=0, leading=24)
        header_sub    = make_style('HSub',    fontSize=9,  textColor=colors.HexColor('#93C5FD'), alignment=TA_CENTER, fontName='Helvetica', spaceAfter=0, spaceBefore=0)
        meta_left     = make_style('MetaL',   fontSize=9,  textColor=NAVY,     fontName='Helvetica-Bold')
        meta_right    = make_style('MetaR',   fontSize=8,  textColor=GREY,     fontName='Helvetica', alignment=TA_RIGHT)
        footer_style  = make_style('Footer',  fontSize=7,  textColor=GREY,     fontName='Helvetica', alignment=TA_CENTER, leading=10)
        sum_label     = make_style('SumLbl',  fontSize=9,  textColor=DARK,     fontName='Helvetica-Bold')
        sum_val       = make_style('SumVal',  fontSize=9,  textColor=DARK,     fontName='Helvetica',  alignment=TA_RIGHT)
        tbl_hdr       = make_style('TblHdr',  fontSize=8,  textColor=WHITE,    fontName='Helvetica-Bold', alignment=TA_CENTER)
        cell_center   = make_style('CellC',   fontSize=7,  textColor=DARK,     fontName='Helvetica', alignment=TA_CENTER)
        cell_right    = make_style('CellR',   fontSize=8,  textColor=DARK,     fontName='Helvetica-Bold', alignment=TA_RIGHT)
        cell_left     = make_style('CellL',   fontSize=7,  textColor=DARK,     fontName='Helvetica')
        cell_grey     = make_style('CellG',   fontSize=7,  textColor=GREY,     fontName='Helvetica', alignment=TA_CENTER)
        income_type   = make_style('IncT',    fontSize=7,  textColor=INCOME_TXT,  fontName='Helvetica-Bold', alignment=TA_CENTER)
        expense_type  = make_style('ExpT',    fontSize=7,  textColor=EXPENSE_TXT, fontName='Helvetica-Bold', alignment=TA_CENTER)

        story = []

        # ══════════════════════════════════════════════════════════════════
        # HEADER BANNER  (navy block with title)
        # ══════════════════════════════════════════════════════════════════
        banner_data = [
            [Paragraph("🏨  HOTEL LODGE", header_title)],
            [Paragraph("OFFICIAL ACCOUNT STATEMENT", header_sub)],
        ]
        banner = Table(banner_data, colWidths=[CONTENT_W])
        banner.setStyle(TableStyle([
            ('BACKGROUND',    (0, 0), (-1, -1), NAVY),
            ('TOPPADDING',    (0, 0), (-1, -1), 8),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
            ('LEFTPADDING',   (0, 0), (-1, -1), 10),
            ('RIGHTPADDING',  (0, 0), (-1, -1), 10),
            ('LINEBELOW',     (0, 1), (-1, 1),  2, BLUE),
        ]))
        story.append(banner)
        story.append(Spacer(1, 3 * mm))

        # ══════════════════════════════════════════════════════════════════
        # META ROW  (holder + generated date)
        # ══════════════════════════════════════════════════════════════════
        meta_data = [[
            Paragraph(f"Account Holder: <b>{user_name}</b>", meta_left),
            Paragraph(f"Generated: {generated_at}", meta_right),
        ]]
        meta_tbl = Table(meta_data, colWidths=[CONTENT_W * 0.6, CONTENT_W * 0.4])
        meta_tbl.setStyle(TableStyle([
            ('BACKGROUND',    (0, 0), (-1, -1), colors.HexColor('#F0F9FF')),
            ('TOPPADDING',    (0, 0), (-1, -1), 5),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
            ('LEFTPADDING',   (0, 0), (-1, -1), 8),
            ('RIGHTPADDING',  (0, 0), (-1, -1), 8),
            ('LINEBELOW',     (0, 0), (-1, 0), 0.5, DIVIDER),
        ]))
        story.append(meta_tbl)
        story.append(Spacer(1, 4 * mm))

        # ══════════════════════════════════════════════════════════════════
        # SUMMARY CARDS  (Income | Expenses | Net Balance)
        # ══════════════════════════════════════════════════════════════════
        W3 = CONTENT_W / 3

        def card_para(label, value, label_color, value_color):
            return Paragraph(
                f'<font color="#{label_color}" size="8"><b>{label}</b></font><br/>'
                f'<font color="#{value_color}" size="11"><b>Rs. {value:,.2f}</b></font>',
                make_style('Card', alignment=TA_CENTER, leading=16, spaceAfter=0)
            )

        bal_sign = "+" if net_balance >= 0 else ""
        summary_cards = [[
            card_para("▲  TOTAL INCOME",   float(total_income),   "065F46", "065F46"),
            card_para("▼  TOTAL EXPENSES", float(total_expenses), "991B1B", "991B1B"),
            Paragraph(
                f'<font color="#1D4ED8" size="8"><b>◆  NET BALANCE</b></font><br/>'
                f'<font color="{"#065F46" if net_balance>=0 else "#991B1B"}" size="11">'
                f'<b>Rs. {bal_sign}{float(net_balance):,.2f}</b></font>',
                make_style('CardB', alignment=TA_CENTER, leading=16, spaceAfter=0)
            ),
        ]]
        card_tbl = Table(summary_cards, colWidths=[W3, W3, W3])
        card_tbl.setStyle(TableStyle([
            ('BACKGROUND',    (0, 0), (0, 0), INCOME_BG),
            ('BACKGROUND',    (1, 0), (1, 0), EXPENSE_BG),
            ('BACKGROUND',    (2, 0), (2, 0), LIGHT_BLUE),
            ('TOPPADDING',    (0, 0), (-1, -1), 10),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 10),
            ('LEFTPADDING',   (0, 0), (-1, -1), 6),
            ('RIGHTPADDING',  (0, 0), (-1, -1), 6),
            ('BOX',           (0, 0), (-1, -1), 0.8, BLUE),
            ('INNERGRID',     (0, 0), (-1, -1), 0.5, DIVIDER),
            ('LINEBEFORE',    (2, 0), (2, 0),   1.5, BLUE),
        ]))
        story.append(card_tbl)
        story.append(Spacer(1, 5 * mm))

        # ══════════════════════════════════════════════════════════════════
        # TRANSACTION TABLE
        # ══════════════════════════════════════════════════════════════════
        # Column widths for A4 portrait content width (~174mm)
        # #(8) | Date(24) | Type(18) | Category(28) | Amount(24) | Recorded(30) | Remarks(42)
        col_widths = [8*mm, 24*mm, 18*mm, 28*mm, 24*mm, 30*mm, None]
        # last col fills remaining space
        used = sum(w for w in col_widths if w)
        col_widths[-1] = CONTENT_W - used

        headers = ["#", "Date (BS)", "Type", "Category", "Amount (Rs.)", "Recorded At", "Remarks"]
        tbl_data = [[Paragraph(h, tbl_hdr) for h in headers]]

        style_cmds = [
            # Header row
            ('BACKGROUND',    (0, 0), (-1, 0),  NAVY),
            ('TEXTCOLOR',     (0, 0), (-1, 0),  WHITE),
            ('FONTNAME',      (0, 0), (-1, 0),  'Helvetica-Bold'),
            ('FONTSIZE',      (0, 0), (-1, 0),  8),
            ('ALIGN',         (0, 0), (-1, 0),  'CENTER'),
            ('VALIGN',        (0, 0), (-1, 0),  'MIDDLE'),
            ('TOPPADDING',    (0, 0), (-1, 0),  6),
            ('BOTTOMPADDING', (0, 0), (-1, 0),  6),
            # All data rows base
            ('FONTNAME',      (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE',      (0, 1), (-1, -1), 7),
            ('VALIGN',        (0, 1), (-1, -1), 'MIDDLE'),
            ('TOPPADDING',    (0, 1), (-1, -1), 4),
            ('BOTTOMPADDING', (0, 1), (-1, -1), 4),
            ('LEFTPADDING',   (0, 0), (-1, -1), 4),
            ('RIGHTPADDING',  (0, 0), (-1, -1), 4),
            # Grid
            ('BOX',           (0, 0), (-1, -1), 0.8, NAVY),
            ('INNERGRID',     (0, 0), (-1, -1), 0.3, colors.HexColor('#E5E7EB')),
            # Column alignments
            ('ALIGN',         (0, 1), (0, -1),  'CENTER'),   # #
            ('ALIGN',         (1, 1), (1, -1),  'CENTER'),   # date
            ('ALIGN',         (2, 1), (2, -1),  'CENTER'),   # type
            ('ALIGN',         (3, 1), (3, -1),  'CENTER'),   # category
            ('ALIGN',         (4, 1), (4, -1),  'RIGHT'),    # amount
            ('ALIGN',         (5, 1), (5, -1),  'CENTER'),   # recorded
            ('ALIGN',         (6, 1), (6, -1),  'LEFT'),     # remarks
            # Amount column bold
            ('FONTNAME',      (4, 1), (4, -1),  'Helvetica-Bold'),
            ('FONTSIZE',      (4, 1), (4, -1),  8),
        ]

        for idx, r in enumerate(records, 1):
            is_income = r['type'] == 'Income'
            row_bg    = INCOME_BG  if is_income else EXPENSE_BG
            alt_bg    = colors.HexColor('#F0FDF4') if is_income else colors.HexColor('#FFF5F5')
            type_style = income_type if is_income else expense_type
            txt_color  = INCOME_TXT if is_income else EXPENSE_TXT
            bg = row_bg if idx % 2 == 1 else alt_bg

            amount_str = f"{r['amount']:,.2f}"
            recorded   = str(r['created_at'])[:16]
            remarks_txt = r['remarks'][:55] + ('…' if len(r['remarks']) > 55 else '')

            row_data = [
                Paragraph(str(idx), cell_grey),
                Paragraph(r['nepali_date'], cell_center),
                Paragraph(r['type'], type_style),
                Paragraph(r['category'], cell_center),
                Paragraph(amount_str, make_style(f'Amt{idx}', fontSize=8,
                          textColor=txt_color, fontName='Helvetica-Bold', alignment=TA_RIGHT)),
                Paragraph(recorded, cell_grey),
                Paragraph(remarks_txt, cell_left),
            ]
            tbl_data.append(row_data)
            data_row_idx = idx  # 1-based offset from header
            style_cmds.append(('BACKGROUND', (0, data_row_idx), (-1, data_row_idx), bg))
            # Bold amount text color
            style_cmds.append(('TEXTCOLOR', (4, data_row_idx), (4, data_row_idx), txt_color))

        tx_table = Table(tbl_data, colWidths=col_widths, repeatRows=1, splitByRow=True)
        tx_table.setStyle(TableStyle(style_cmds))
        story.append(tx_table)
        story.append(Spacer(1, 5 * mm))

        # ══════════════════════════════════════════════════════════════════
        # FOOTER TOTALS BAR
        # ══════════════════════════════════════════════════════════════════
        bal_sign_str = "+" if net_balance >= 0 else ""
        footer_data = [[
            Paragraph(f"Total Transactions: <b>{total_records}</b>", meta_left),
            Paragraph(f"Income: <b>Rs. {float(total_income):,.2f}</b>  &nbsp;&nbsp;  "
                      f"Expenses: <b>Rs. {float(total_expenses):,.2f}</b>  &nbsp;&nbsp;  "
                      f"Net: <b>Rs. {bal_sign_str}{float(net_balance):,.2f}</b>",
                      make_style('FootR', fontSize=9, fontName='Helvetica',
                                 textColor=NAVY, alignment=TA_RIGHT)),
        ]]
        footer_tbl = Table(footer_data, colWidths=[CONTENT_W * 0.4, CONTENT_W * 0.6])
        footer_tbl.setStyle(TableStyle([
            ('BACKGROUND',    (0, 0), (-1, -1), LIGHT_BLUE),
            ('TOPPADDING',    (0, 0), (-1, -1), 7),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 7),
            ('LEFTPADDING',   (0, 0), (-1, -1), 8),
            ('RIGHTPADDING',  (0, 0), (-1, -1), 8),
            ('BOX',           (0, 0), (-1, -1), 1, NAVY),
        ]))
        story.append(footer_tbl)
        story.append(Spacer(1, 4 * mm))

        # ── Disclaimer line ────────────────────────────────────────────────
        story.append(HRFlowable(width=CONTENT_W, color=DIVIDER, thickness=0.5))
        story.append(Spacer(1, 2 * mm))
        story.append(Paragraph(
            "This is a system-generated statement. No signature is required. "
            "For queries, contact the hotel management.",
            footer_style
        ))

        # ── Page numbering via onFirstPage / onLaterPages canvas callback ──
        def add_page_number(canvas, doc):
            canvas.saveState()
            canvas.setFont('Helvetica', 7)
            canvas.setFillColor(GREY)
            page_num  = canvas.getPageNumber()
            text = f"Page {page_num}  |  Hotel Lodge — Statement"
            canvas.drawCentredString(PAGE_W / 2, 10 * mm, text)
            # Top right watermark
            canvas.setFont('Helvetica', 7)
            canvas.setFillColor(colors.HexColor('#BFDBFE'))
            canvas.drawRightString(PAGE_W - MARGIN, PAGE_H - 8 * mm, "CONFIDENTIAL")
            canvas.restoreState()

        doc.build(story, onFirstPage=add_page_number, onLaterPages=add_page_number)
        buffer.seek(0)

        response = HttpResponse(buffer, content_type="application/pdf")
        response["Content-Disposition"] = 'attachment; filename="hotel_lodge_statement.pdf"'
        return response
